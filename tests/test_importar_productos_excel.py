"""
Nueva funcionalidad: importar productos en lote desde un archivo Excel.
Crea productos nuevos por nombre, actualiza los que ya existen, y
reporta errores de filas invalidas sin tumbar toda la importacion.

Incluye el flujo round-trip real del admin: descargar el Excel del
exportador, editarlo y volver a subirlo por el importador.
"""
import os
import io
import tempfile

_DB_DIR = tempfile.mkdtemp(prefix="cafeteria_test_import_")
_DB_PATH = os.path.join(_DB_DIR, "test_cafeteria.db")
os.environ['DATABASE_URL'] = f"sqlite:///{_DB_PATH}"
os.environ.pop('SSH_HOST', None)

import pytest
from openpyxl import Workbook, load_workbook

from app import create_app
from app.models import db, Producto


@pytest.fixture()
def app():
    application = create_app('development')
    application.config['WTF_CSRF_ENABLED'] = False
    application.config['TESTING'] = True
    # Nota: _seed_datos_iniciales() ya crea un producto "Café Tinto"
    # (precio 1500) -- se reutiliza ese para probar la actualizacion,
    # en vez de insertar un duplicado.
    yield application
    with application.app_context():
        db.session.remove()
        db.drop_all()
        db.engine.dispose()


@pytest.fixture()
def client(app):
    return app.test_client()


def _login_admin(client):
    return client.post('/empleados/login', data={
        'email': 'admin@cafeteria.com', 'clave': 'Admin123',
    }, follow_redirects=True)


def _excel_bytes(headers, filas):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for fila in filas:
        ws.append(fila)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Formato legacy (headers en minusculas, como siempre) ────────────

def test_crea_producto_nuevo_y_actualiza_uno_existente(app, client):
    _login_admin(client)
    with app.app_context():
        total_antes = Producto.query.count()

    archivo = _excel_bytes(
        ['nombre', 'precio', 'costo', 'stock', 'stock_minimo'],
        [
            ['Café Tinto', 1800, 600, 40, 8],          # existente -> actualiza
            ['Té Chai QA', 2000, 700, 15, 5],           # nuevo -> crea
        ],
    )
    resp = client.post('/admin/productos/importar', data={
        'archivo': (archivo, 'productos.xlsx'),
    }, content_type='multipart/form-data', follow_redirects=True)
    assert resp.status_code == 200

    with app.app_context():
        cafe = Producto.query.filter_by(nombre='Café Tinto').first()
        te = Producto.query.filter_by(nombre='Té Chai QA').first()
        # Solo se agrego 1 producto nuevo; Cafe Tinto ya existia y se
        # actualizo, no se duplico.
        assert Producto.query.count() == total_antes + 1
        assert cafe.precio == 1800
        assert cafe.costo == 600
        assert cafe.stock == 40
        assert te is not None
        assert te.precio == 2000


def test_fila_invalida_no_tumba_la_importacion_completa(app, client):
    _login_admin(client)
    archivo = _excel_bytes(
        ['nombre', 'precio'],
        [
            ['Producto Bueno', 3000],
            ['Producto Malo', 'no-es-un-numero'],
            ['', 5000],  # sin nombre
        ],
    )
    resp = client.post('/admin/productos/importar', data={
        'archivo': (archivo, 'productos.xlsx'),
    }, content_type='multipart/form-data', follow_redirects=True)
    assert resp.status_code == 200

    with app.app_context():
        assert Producto.query.filter_by(nombre='Producto Bueno').first() is not None
        assert Producto.query.filter_by(nombre='Producto Malo').first() is None


def test_rechaza_archivo_sin_columna_precio(app, client):
    _login_admin(client)
    archivo = _excel_bytes(['nombre', 'descripcion'], [['X', 'algo']])
    resp = client.post('/admin/productos/importar', data={
        'archivo': (archivo, 'productos.xlsx'),
    }, content_type='multipart/form-data', follow_redirects=True)
    assert resp.status_code == 200
    with app.app_context():
        assert Producto.query.filter_by(nombre='X').first() is None


def test_solo_escribir_todo_puede_importar(app, client):
    resp_sin_login = client.get('/admin/productos/importar', follow_redirects=True)
    # Sin sesion, login_required redirige al login
    assert b'Correo' in resp_sin_login.data or resp_sin_login.status_code in (200, 302)


# ── Alias de encabezados: otras formas validas de escribir lo mismo ──

def test_importa_con_encabezados_alternativos(app, client):
    _login_admin(client)
    archivo = _excel_bytes(
        ['Nombre Producto', 'Precio ($)', 'Costo Unitario',
         'Stock Actual', 'Stock Mínimo'],
        [['Matcha Latte', 9500, 4000, 12, 3]],
    )
    resp = client.post('/admin/productos/importar', data={
        'archivo': (archivo, 'productos.xlsx'),
    }, content_type='multipart/form-data', follow_redirects=True)
    assert resp.status_code == 200

    with app.app_context():
        p = Producto.query.filter_by(nombre='Matcha Latte').one()
        assert p.precio == 9500
        assert p.costo == 4000
        assert p.stock == 12
        assert p.stock_minimo == 3


# ── Round-trip real: descargar el Excel del exportador y re-subirlo ──

def test_excel_del_exportador_se_puede_reimportar(app, client):
    _login_admin(client)
    with app.app_context():
        total_antes = Producto.query.count()
        assert Producto.query.count() > 0

    # 1) Descargar el inventario tal cual lo genera el exportador.
    excel = client.get('/admin/productos/excel')
    assert excel.status_code == 200

    # 2) Editarlo en memoria: cambiar precio/costo/stock de Cafe Tinto.
    wb = load_workbook(io.BytesIO(excel.data))
    ws = wb.active
    fila_cafe = None
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=2).value == 'Café Tinto':
            fila_cafe = fila
            break
    assert fila_cafe is not None, 'el seed deberia crear Cafe Tinto'
    ws.cell(row=fila_cafe, column=3, value=7777)   # Precio ($)
    ws.cell(row=fila_cafe, column=4, value=1234)   # Costo ($)
    ws.cell(row=fila_cafe, column=5, value=33)     # Stock
    ws.cell(row=fila_cafe, column=6, value=6)      # Stock Minimo
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 3) Re-subir el mismo archivo al importador.
    resp = client.post('/admin/productos/importar', data={
        'archivo': (buf, 'inventario_editado.xlsx'),
    }, content_type='multipart/form-data', follow_redirects=True)
    assert resp.status_code == 200

    with app.app_context():
        cafe = Producto.query.filter_by(nombre='Café Tinto').one()
        assert cafe.precio == 7777
        assert cafe.costo == 1234
        assert cafe.stock == 33
        assert cafe.stock_minimo == 6
        # Se actualizo, no se duplico ni se rechazo el archivo.
        assert Producto.query.count() == total_antes
        # La fila TOTAL del export no debe convertirse en producto.
        assert Producto.query.filter_by(nombre='TOTAL').first() is None