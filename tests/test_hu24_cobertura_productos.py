"""
HU-24: cierra huecos de cobertura en blueprints/admin/productos.py que
no tenian ninguna prueba -- exportacion a Excel del inventario,
endpoint de imagenes de biblioteca, y ramas de validacion de la
importacion masiva (archivo faltante, extension invalida, columnas
no reconocidas).
"""
import os
import io
import tempfile

_DB_DIR = tempfile.mkdtemp(prefix="cafeteria_test_hu24_prod_")
_DB_PATH = os.path.join(_DB_DIR, "test_cafeteria.db")
os.environ['DATABASE_URL'] = f"sqlite:///{_DB_PATH}"
os.environ.pop('SSH_HOST', None)

import pytest
from openpyxl import Workbook, load_workbook

from app import create_app
from app.models import db


@pytest.fixture()
def app():
    application = create_app('development')
    application.config['WTF_CSRF_ENABLED'] = False
    application.config['TESTING'] = True
    yield application
    with application.app_context():
        db.session.remove()
        db.drop_all()
        db.engine.dispose()


@pytest.fixture()
def client(app):
    return app.test_client()


def _login_admin(client):
    return client.post('/empleados/login', data={
        'email': 'admin@cafeteria.com', 'clave': 'Admin123',
    }, follow_redirects=True)


def test_exportar_inventario_a_excel(client):
    _login_admin(client)
    resp = client.get('/admin/productos/excel')
    assert resp.status_code == 200
    assert resp.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    encabezados = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert encabezados == ['ID', 'Producto', 'Precio ($)', 'Stock', 'Estado']
    # el seed ya crea productos -- debe haber al menos una fila de datos
    assert ws.max_row > 1


def test_endpoint_de_imagenes_de_biblioteca_devuelve_lista_json(client):
    _login_admin(client)
    resp = client.get('/admin/productos/imagenes')
    assert resp.status_code == 200
    assert isinstance(resp.get_json(), list)


def test_importar_sin_archivo_muestra_error(client):
    _login_admin(client)
    resp = client.post('/admin/productos/importar', data={}, follow_redirects=True)
    assert 'Selecciona un archivo Excel' in resp.get_data(as_text=True)


def test_importar_con_extension_invalida_se_rechaza(client):
    _login_admin(client)
    archivo = (io.BytesIO(b'no es un excel'), 'productos.txt')
    resp = client.post('/admin/productos/importar',
                        data={'archivo': archivo}, content_type='multipart/form-data',
                        follow_redirects=True)
    assert 'debe ser .xlsx' in resp.get_data(as_text=True)


def test_importar_avisa_columnas_no_reconocidas(client):
    _login_admin(client)
    wb = Workbook()
    ws = wb.active
    ws.append(['nombre', 'precio', 'columna_rara'])
    ws.append(['Producto columna extra', 3000, 'valor sin uso'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post('/admin/productos/importar',
                        data={'archivo': (buf, 'productos.xlsx')},
                        content_type='multipart/form-data', follow_redirects=True)
    assert 'no reconocidas e ignoradas' in resp.get_data(as_text=True)
    assert 'columna_rara' in resp.get_data(as_text=True)
