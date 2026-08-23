from flask import render_template, redirect, url_for, request, flash, jsonify, send_file, current_app
from flask_login import login_required, current_user
from PIL import Image
import os
import shutil
import unicodedata
import uuid
from ...models import db, expr_fecha, Producto, BajaInventario, ajustar_stock, registrar_auditoria
from ...models.producto import CATEGORIAS
from ..permisos import requiere_permiso, requiere_ver_pagina
from ...utils import ahora_bogota, hoy_bogota
from datetime import datetime, timedelta
from . import admin_bp


ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_IMAGE_DIMENSION = 4000  # px; evita decompression bombs


def _allowed_image(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _es_imagen_valida(file_storage):
    """Verifica que el archivo sea realmente una imagen (no solo la
    extension) y que sus dimensiones no sean excesivas."""
    try:
        file_storage.stream.seek(0)
        Image.open(file_storage.stream).verify()
    except Exception:
        return False
    try:
        file_storage.stream.seek(0)
        ancho, alto = Image.open(file_storage.stream).size
    except Exception:
        return False
    finally:
        file_storage.stream.seek(0)
    return ancho <= MAX_IMAGE_DIMENSION and alto <= MAX_IMAGE_DIMENSION


def _library_dir():
    return os.path.join(current_app.static_folder, 'imagenes')


def _save_image(file_storage):
    """Guarda imagen en static/productos/ y copia a static/imagenes/ (biblioteca)."""
    if not file_storage or not file_storage.filename:
        return None
    if not _allowed_image(file_storage.filename):
        return None
    if not _es_imagen_valida(file_storage):
        return None
    ext = file_storage.filename.rsplit('.', 1)[1].lower()
    filename = f"{uuid.uuid4().hex[:12]}.{ext}"
    upload_dir = current_app.config['UPLOAD_FOLDER']
    os.makedirs(upload_dir, exist_ok=True)
    file_storage.save(os.path.join(upload_dir, filename))
    # Copiar a biblioteca para reutilización futura
    lib_dir = _library_dir()
    os.makedirs(lib_dir, exist_ok=True)
    shutil.copy2(os.path.join(upload_dir, filename), os.path.join(lib_dir, filename))
    return filename


def _delete_image(filename):
    """Elimina imagen de static/productos/ (NO de la biblioteca)."""
    if not filename:
        return
    path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(path):
        os.remove(path)


def _get_library_images():
    """Retorna lista de imágenes disponibles en la biblioteca."""
    lib_dir = _library_dir()
    if not os.path.exists(lib_dir):
        return []
    allowed = ALLOWED_EXTENSIONS
    return sorted([
        f for f in os.listdir(lib_dir)
        if '.' in f and f.rsplit('.', 1)[1].lower() in allowed and f != '.gitkeep'
    ])


@admin_bp.route('/productos')
@login_required
@requiere_ver_pagina('productos')
def productos():
    # ?inactivos=1 muestra los dados de baja logicamente.
    ver_inactivos = request.args.get('inactivos') == '1'
    query = Producto.query.order_by(Producto.idproducto)
    prods = query.filter(Producto.activo.is_(False)).all() if ver_inactivos \
        else query.filter(Producto.activo.is_(True)).all()
    imagenes = _get_library_images()
    bajas = (BajaInventario.query
             .options(db.joinedload(BajaInventario.producto))
             .order_by(BajaInventario.idbaja.desc())
             .limit(15)
             .all())
    hoy = hoy_bogota()
    semana_lunes = (hoy - timedelta(days=hoy.weekday())).strftime('%Y-%m-%d')
    semana_viernes = ((hoy - timedelta(days=hoy.weekday()))
                      + timedelta(days=4)).strftime('%Y-%m-%d')
    # Sin categoria no aparecen en "Ordenar", solo en el catalogo plano.
    sin_categoria_count = Producto.query.filter(
        Producto.activo.is_(True), Producto.categoria.is_(None)
    ).count()
    return render_template('admin/productos.html', productos=prods,
                           imagenes_biblioteca=imagenes, bajas_recientes=bajas,
                           ver_inactivos=ver_inactivos, categorias_disponibles=CATEGORIAS,
                           semana_lunes=semana_lunes, semana_viernes=semana_viernes,
                           sin_categoria_count=sin_categoria_count)


@admin_bp.route('/productos/nuevo', methods=['POST'])
@login_required
@requiere_permiso('escribir_todo')
def producto_nuevo():
    nombre = request.form.get('nombre', '').strip()
    if not nombre:
        flash('El nombre es obligatorio.', 'danger')
        return redirect(url_for('admin_panel.productos'))
    try:
        precio = int(request.form.get('precio', 0))
        stock = int(request.form.get('stock', 0))
    except (ValueError, TypeError):
        flash('Precio y stock deben ser valores numericos.', 'danger')
        return redirect(url_for('admin_panel.productos'))
    try:
        stock_minimo = int(request.form.get('stock_minimo', 10))
        if stock_minimo <= 0:
            stock_minimo = 10
    except (ValueError, TypeError):
        stock_minimo = 10
    try:
        costo = int(request.form.get('costo', 0))
    except (ValueError, TypeError):
        costo = 0
    # El min="0" del formulario es solo del lado del cliente.
    if precio < 0 or costo < 0 or stock < 0:
        flash('Precio, costo y stock no pueden ser negativos.', 'danger')
        return redirect(url_for('admin_panel.productos'))
    imagen = None
    new_image = request.files.get('imagen')
    if new_image and new_image.filename:
        imagen = _save_image(new_image)
    else:
        lib_image = request.form.get('imagen_biblioteca', '').strip()
        # Solo se permite copiar un nombre que realmente este en la
        # biblioteca, para evitar un path traversal via ese campo.
        if lib_image and lib_image in _get_library_images():
            lib_path = os.path.join(_library_dir(), lib_image)
            if os.path.exists(lib_path):
                dest_dir = current_app.config['UPLOAD_FOLDER']
                os.makedirs(dest_dir, exist_ok=True)
                shutil.copy2(lib_path, os.path.join(dest_dir, lib_image))
                imagen = lib_image
    es_especial = request.form.get('es_especial') == '1'
    categoria = request.form.get('categoria', '').strip() or None
    subcategoria = request.form.get('subcategoria', '').strip() or None
    descripcion = request.form.get('descripcion', '').strip() or None
    prod = Producto(nombre=nombre, precio=precio, stock=stock, imagen=imagen,
                    stock_minimo=stock_minimo, costo=costo, es_especial=es_especial,
                    categoria=categoria, subcategoria=subcategoria, descripcion=descripcion)
    db.session.add(prod)
    db.session.commit()
    registrar_auditoria(current_user.email, 'crear_producto', f'producto:{prod.idproducto}',
                         f'precio={precio}, costo={costo}')
    flash(f'Producto "{nombre}" creado correctamente.', 'success')
    return redirect(url_for('admin_panel.productos'))


@admin_bp.route('/productos/editar/<int:idproducto>', methods=['POST'])
@login_required
@requiere_permiso('escribir_todo')
def producto_editar(idproducto):
    prod = Producto.query.get_or_404(idproducto)
    precio_anterior = prod.precio
    prod.nombre = request.form.get('nombre', prod.nombre).strip()
    try:
        nuevo_precio = int(request.form.get('precio', prod.precio))
        nuevo_stock  = int(request.form.get('stock', prod.stock))
    except (ValueError, TypeError):
        flash('Precio y stock deben ser valores numericos.', 'danger')
        return redirect(url_for('admin_panel.productos'))
    try:
        nuevo_costo = int(request.form.get('costo', prod.costo))
    except (ValueError, TypeError):
        nuevo_costo = prod.costo
    if nuevo_precio < 0 or nuevo_stock < 0 or nuevo_costo < 0:
        flash('Precio, costo y stock no pueden ser negativos.', 'danger')
        return redirect(url_for('admin_panel.productos'))
    prod.precio = nuevo_precio
    prod.stock  = nuevo_stock
    prod.costo  = nuevo_costo
    prod.es_especial = request.form.get('es_especial') == '1'
    prod.categoria = request.form.get('categoria', '').strip() or None
    prod.subcategoria = request.form.get('subcategoria', '').strip() or None
    prod.descripcion = request.form.get('descripcion', '').strip() or None
    try:
        sm = int(request.form.get('stock_minimo', prod.stock_minimo))
        if sm > 0:
            prod.stock_minimo = sm
    except (ValueError, TypeError):
        pass
    new_image = request.files.get('imagen')
    if new_image and new_image.filename:
        _delete_image(prod.imagen)
        prod.imagen = _save_image(new_image)
    else:
        lib_image = request.form.get('imagen_biblioteca', '').strip()
        if lib_image and lib_image in _get_library_images():
            lib_path = os.path.join(_library_dir(), lib_image)
            if os.path.exists(lib_path):
                dest_dir = current_app.config['UPLOAD_FOLDER']
                os.makedirs(dest_dir, exist_ok=True)
                shutil.copy2(lib_path, os.path.join(dest_dir, lib_image))
                prod.imagen = lib_image
    db.session.commit()
    detalle = f'precio {precio_anterior}->{nuevo_precio}' if precio_anterior != nuevo_precio else None
    registrar_auditoria(current_user.email, 'editar_producto', f'producto:{idproducto}', detalle)
    flash(f'Producto "{prod.nombre}" actualizado.', 'success')
    return redirect(url_for('admin_panel.productos'))


@admin_bp.route('/productos/eliminar/<int:idproducto>', methods=['POST'])
@login_required
@requiere_permiso('escribir_todo')
def producto_eliminar(idproducto):
    prod = Producto.query.get_or_404(idproducto)
    nombre_prod = prod.nombre
    # Si tiene historial, no se borra fisicamente: se marca inactivo
    # para preservar ventas/compras/reportes que dependen de el.
    tiene_historial = bool(prod.detalles_venta or prod.detalles_compra or prod.bajas or prod.reportes)
    if tiene_historial:
        prod.activo = False
        db.session.commit()
        registrar_auditoria(current_user.email, 'desactivar_producto', f'producto:{idproducto}', nombre_prod)
        flash(f'Producto "{nombre_prod}" tiene historial asociado y fue marcado como INACTIVO (el historial se conserva).', 'warning')
    else:
        _delete_image(prod.imagen)
        db.session.delete(prod)
        db.session.commit()
        registrar_auditoria(current_user.email, 'eliminar_producto', f'producto:{idproducto}', nombre_prod)
        flash(f'Producto "{nombre_prod}" eliminado.', 'warning')
    return redirect(url_for('admin_panel.productos'))


@admin_bp.route('/productos/reactivar/<int:idproducto>', methods=['POST'])
@login_required
@requiere_permiso('escribir_todo')
def producto_reactivar(idproducto):
    prod = Producto.query.get_or_404(idproducto)
    prod.activo = True
    db.session.commit()
    registrar_auditoria(current_user.email, 'reactivar_producto', f'producto:{idproducto}')
    flash(f'Producto "{prod.nombre}" reactivado.', 'success')
    return redirect(url_for('admin_panel.productos', inactivos=1))


@admin_bp.route('/productos/baja/<int:idproducto>', methods=['POST'])
@login_required
@requiere_permiso('escribir_todo')
def producto_baja(idproducto):
    prod = Producto.query.get_or_404(idproducto)
    try:
        cantidad = int(request.form.get('cantidad', 0))
    except (ValueError, TypeError):
        flash('La cantidad debe ser un valor numerico.', 'danger')
        return redirect(url_for('admin_panel.productos'))
    motivo = request.form.get('motivo', '').strip()

    if cantidad <= 0 or not motivo:
        flash('La cantidad y el motivo son obligatorios.', 'danger')
        return redirect(url_for('admin_panel.productos'))

    # UPDATE condicional atomico: evita que dos bajas concurrentes dejen el stock negativo.
    if not ajustar_stock(idproducto, -cantidad):
        flash(f'No puedes dar de baja más unidades de las que hay en stock ({prod.stock}).', 'danger')
        return redirect(url_for('admin_panel.productos'))

    categoria = motivo if motivo in ('Vencido', 'Dañado', 'Otro') else 'Otro'
    uid = current_user.get_id()
    baja = BajaInventario(
        idproducto=idproducto, cantidad=cantidad, motivo=motivo, categoria=categoria,
        usuario_documento=current_user.documento if uid.startswith('admin:') else current_user.docpersonal,
        usuario_tipo='admin' if uid.startswith('admin:') else 'personal',
    )
    db.session.add(baja)
    db.session.commit()
    registrar_auditoria(current_user.email, 'baja_inventario', f'producto:{idproducto}',
                         f'cantidad={cantidad}, motivo={motivo}')

    flash(f'Se han dado de baja {cantidad} unidades de "{prod.nombre}" por motivo: {motivo}.', 'success')
    return redirect(url_for('admin_panel.productos'))


@admin_bp.route('/productos/imagenes')
@login_required
@requiere_permiso('escribir_todo')
def productos_imagenes():
    """Retorna JSON con las imágenes disponibles en la biblioteca."""
    imagenes = _get_library_images()
    return jsonify(imagenes)


def _parsear_valor_numerico(valor, minimo=0, permitir_none=False):
    """Convierte una celda de Excel a entero >= minimo. Acepta int/float
    (Excel guarda numeros como float) y strings numericos. Devuelve
    None si la celda esta vacia y permitir_none=True, o lanza ValueError."""
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        if permitir_none:
            return None
        raise ValueError("valor vacio")
    numero = int(float(valor))
    if numero < minimo:
        raise ValueError(f"debe ser >= {minimo}")
    return numero


def _normalizar(texto):
    """Minusculas, sin tildes, sin espacios/guiones extra, para poder
    comparar encabezados escritos de formas distintas ('Stock Mínimo',
    'stock_minimo', 'Stock Minimo' deben ser la misma columna)."""
    texto = str(texto or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    texto = texto.replace('($)', '').replace('(', '').replace(')', '')
    texto = texto.strip().replace(' ', '_').replace('-', '_')
    while '__' in texto:
        texto = texto.replace('__', '_')
    return texto.strip('_')


ALIAS_COLUMNAS = {
    'producto': 'nombre',
    'nombre': 'nombre',
    'nombre_producto': 'nombre',
    'precio': 'precio',
    'precio_venta': 'precio',
    'costo': 'costo',
    'costo_unitario': 'costo',
    'stock': 'stock',
    'stock_actual': 'stock',
    'stock_minimo': 'stock_minimo',
    'stock_min': 'stock_minimo',
    'categoria': 'categoria',       # 'Categoría' normaliza igual (sin tildes)
    'subcategoria': 'subcategoria',
}


@admin_bp.route('/productos/importar', methods=['GET', 'POST'])
@login_required
@requiere_permiso('escribir_todo')
def productos_importar():
    """Carga masiva de productos desde un Excel. Columnas reconocidas
    por nombre de encabezado (no por posicion), con alias flexibles para
    aceptar tambien el formato del exportador ('Producto', 'Precio ($)',
    'Stock Mínimo'...). Si el nombre ya existe (comparacion insensible
    a mayusculas), se ACTUALIZA ese producto; si no, se crea uno nuevo."""
    if request.method == 'GET':
        return render_template('admin/productos_importar.html')

    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        flash('Selecciona un archivo Excel (.xlsx).', 'danger')
        return redirect(url_for('admin_panel.productos_importar'))
    if not archivo.filename.lower().endswith('.xlsx'):
        flash('El archivo debe ser .xlsx.', 'danger')
        return redirect(url_for('admin_panel.productos_importar'))

    from openpyxl import load_workbook

    try:
        wb = load_workbook(archivo.stream, read_only=True, data_only=True)
    except Exception:
        flash('No se pudo leer el archivo. Confirma que sea un Excel valido (.xlsx).', 'danger')
        return redirect(url_for('admin_panel.productos_importar'))

    ws = wb.active
    filas = ws.iter_rows(values_only=True)
    try:
        encabezados = next(filas)
    except StopIteration:
        flash('El archivo esta vacio.', 'danger')
        return redirect(url_for('admin_panel.productos_importar'))

    columnas = {}
    for i, v in enumerate(encabezados):
        if v is None:
            continue
        clave = ALIAS_COLUMNAS.get(_normalizar(v))
        if clave and clave not in columnas:  # la primera columna que matchea gana
            columnas[clave] = i
    if 'nombre' not in columnas or 'precio' not in columnas:
        flash("El Excel debe tener al menos las columnas 'nombre' y 'precio' (o 'Producto' y 'Precio ($)').", 'danger')
        return redirect(url_for('admin_panel.productos_importar'))

    productos_por_nombre = {
        p.nombre.strip().lower(): p for p in Producto.query.all()
    }

    creados = 0
    actualizados = 0
    errores = []

    for num_fila, fila in enumerate(filas, start=2):
        if fila is None or all(c is None for c in fila):
            continue
        nombre = str(fila[columnas['nombre']] or '').strip()
        if not nombre:
            errores.append(f'Fila {num_fila}: falta el nombre, se omite.')
            continue
        # El exportador agrega una fila de totales; re-importarla crearia
        # un producto basura "TOTAL".
        if nombre.lower() == 'total':
            continue
        try:
            precio = _parsear_valor_numerico(fila[columnas['precio']], minimo=0)
            costo = _parsear_valor_numerico(fila[columnas['costo']], minimo=0) if 'costo' in columnas else 0
            stock = _parsear_valor_numerico(fila[columnas['stock']], minimo=0) if 'stock' in columnas else 0
            stock_minimo = _parsear_valor_numerico(fila[columnas['stock_minimo']], minimo=1) if 'stock_minimo' in columnas else 10
            if costo is None:
                costo = 0
            if stock is None:
                stock = 0
            if stock_minimo is None:
                stock_minimo = 10
        except ValueError as e:
            errores.append(f'Fila {num_fila} ("{nombre}"): {e}')
            continue

        # Categoria/subcategoria: solo se tocan si la columna viene en el
        # Excel. Una categoria fuera del catalogo NO descarta la fila:
        # se reporta y el producto queda como estaba (o sin categoria).
        nueva_categoria = None
        if 'categoria' in columnas:
            valor = str(fila[columnas['categoria']] or '').strip()
            if valor:
                if valor in CATEGORIAS:
                    nueva_categoria = valor
                else:
                    errores.append(
                        f'Fila {num_fila} ("{nombre}"): categoria "{valor}" no reconocida, '
                        f'se deja sin categoria. Categorias validas: {", ".join(CATEGORIAS)}.'
                    )
        nueva_subcategoria = None
        if 'subcategoria' in columnas:
            valor = str(fila[columnas['subcategoria']] or '').strip()
            if valor:
                nueva_subcategoria = valor

        clave = nombre.lower()
        existente = productos_por_nombre.get(clave)
        if existente:
            existente.precio = precio
            existente.costo = costo
            existente.stock = stock
            existente.stock_minimo = stock_minimo
            if nueva_categoria is not None:
                existente.categoria = nueva_categoria
            if nueva_subcategoria is not None:
                existente.subcategoria = nueva_subcategoria
            actualizados += 1
        else:
            nuevo = Producto(nombre=nombre, precio=precio, costo=costo,
                            stock=stock, stock_minimo=stock_minimo,
                            categoria=nueva_categoria,
                            subcategoria=nueva_subcategoria)
            db.session.add(nuevo)
            productos_por_nombre[clave] = nuevo
            creados += 1

    db.session.commit()

    # columnas ya guarda claves canonicas: las ignoradas se reportan
    # desde los encabezados crudos que no mapearon a nada.
    columnas_ignoradas = {
        str(v) for v in encabezados
        if v is not None and ALIAS_COLUMNAS.get(_normalizar(v)) is None
    }
    resumen = f'Importacion completa: {creados} producto(s) creado(s), {actualizados} actualizado(s).'
    if columnas_ignoradas:
        resumen += f' Columnas no reconocidas e ignoradas: {", ".join(sorted(columnas_ignoradas))}.'
    flash(resumen, 'success' if not errores else 'warning')
    for err in errores[:20]:
        flash(err, 'warning')
    if len(errores) > 20:
        flash(f'... y {len(errores) - 20} error(es) mas, no mostrados.', 'warning')

    return redirect(url_for('admin_panel.productos'))


@admin_bp.route('/productos/excel')
@login_required
@requiere_ver_pagina('productos')
def productos_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    productos = Producto.query.filter(Producto.activo.is_(True)).order_by(Producto.nombre).limit(5000).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ['ID', 'Producto', 'Precio ($)', 'Costo ($)', 'Stock', 'Stock Mínimo',
               'Estado', 'Categoría', 'Subcategoría', 'Descripción', 'Es Especial',
               'Especial Hasta', 'Creado', 'Actualizado', 'Activo']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, p in enumerate(productos, 2):
        ws.cell(row=row_idx, column=1, value=p.idproducto).border = thin_border
        ws.cell(row=row_idx, column=2, value=p.nombre).border = thin_border
        ws.cell(row=row_idx, column=3, value=p.precio).border = thin_border
        ws.cell(row=row_idx, column=4, value=p.costo).border = thin_border
        ws.cell(row=row_idx, column=5, value=p.stock).border = thin_border
        ws.cell(row=row_idx, column=6, value=p.stock_minimo).border = thin_border
        ws.cell(row=row_idx, column=7, value=p.estado).border = thin_border
        ws.cell(row=row_idx, column=8, value=p.categoria or '').border = thin_border
        ws.cell(row=row_idx, column=9, value=p.subcategoria or '').border = thin_border
        ws.cell(row=row_idx, column=10, value=p.descripcion or '').border = thin_border
        ws.cell(row=row_idx, column=11, value='Sí' if p.es_especial else 'No').border = thin_border
        ws.cell(row=row_idx, column=12,
                value=p.especial_hasta.strftime('%d/%m/%Y') if p.especial_hasta else '').border = thin_border
        ws.cell(row=row_idx, column=13,
                value=p.created_at.strftime('%d/%m/%Y %H:%M') if p.created_at else '').border = thin_border
        ws.cell(row=row_idx, column=14,
                value=p.updated_at.strftime('%d/%m/%Y %H:%M') if p.updated_at else '').border = thin_border
        ws.cell(row=row_idx, column=15, value='Sí' if p.activo else 'No').border = thin_border

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 35
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 10

    # Fila de totales sobre lo REALMENTE exportado (el query ya excluye
    # inactivos): valor del inventario a precio y a costo.
    fila_total = len(productos) + 2
    fill_total = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    font_total = Font(bold=True)
    totales = {
        2: 'TOTAL',
        3: sum(p.precio * p.stock for p in productos),
        4: sum(p.costo * p.stock for p in productos),
    }
    for col in range(1, len(headers) + 1):
        celda = ws.cell(row=fila_total, column=col, value=totales.get(col))
        celda.font = font_total
        celda.fill = fill_total
        celda.border = thin_border

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = ahora_bogota().strftime('%Y-%m-%d_%H-%M')
    return send_file(buf, as_attachment=True,
                     download_name=f"inventario_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route('/productos/bajas/excel')
@login_required
@requiere_ver_pagina('productos')
def productos_bajas_excel():
    """Excel de bajas de inventario de la ultima semana (o del rango
    ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD si se pasa)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    # Rango opcional; por defecto los ultimos 7 dias. Un rango invalido
    # nunca debe romper la respuesta: cae al default.
    MAX_FILAS_EXCEL = 5000
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')
    hoy = hoy_bogota()
    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str \
            else hoy - timedelta(days=6)
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy
    except ValueError:
        desde, hasta = hoy - timedelta(days=6), hoy

    bajas = (BajaInventario.query
             .options(db.joinedload(BajaInventario.producto))
             .filter(expr_fecha(BajaInventario.fecha) >= desde,
                     expr_fecha(BajaInventario.fecha) <= hasta)
             .order_by(BajaInventario.idbaja.asc())
             .limit(MAX_FILAS_EXCEL)
             .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Bajas de Inventario"

    headers = ['ID Baja', 'Producto', 'Cantidad', 'Motivo', 'Categoría', 'Fecha', 'Hora', 'Usuario']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, b in enumerate(bajas, 2):
        ws.cell(row=row_idx, column=1, value=b.idbaja).border = thin_border
        ws.cell(row=row_idx, column=2,
                value=b.producto.nombre if b.producto else '').border = thin_border
        ws.cell(row=row_idx, column=3, value=b.cantidad).border = thin_border
        ws.cell(row=row_idx, column=4, value=b.motivo).border = thin_border
        ws.cell(row=row_idx, column=5, value=b.categoria or '').border = thin_border
        ws.cell(row=row_idx, column=6,
                value=b.fecha.strftime('%d/%m/%Y') if b.fecha else '').border = thin_border
        ws.cell(row=row_idx, column=7,
                value=b.fecha.strftime('%H:%M') if b.fecha else '').border = thin_border
        ws.cell(row=row_idx, column=8, value=b.usuario_nombre or '').border = thin_border

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 25

    # Fila de total de unidades dadas de baja (consistente con lo exportado).
    fila_total = len(bajas) + 2
    fill_total = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    font_total = Font(bold=True)
    totales = {
        2: 'TOTAL UNIDADES DE BAJA',
        3: sum(b.cantidad for b in bajas),
    }
    for col in range(1, len(headers) + 1):
        celda = ws.cell(row=fila_total, column=col, value=totales.get(col))
        celda.font = font_total
        celda.fill = fill_total
        celda.border = thin_border

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = ahora_bogota().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f"bajas_inventario_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
