from flask import render_template, redirect, url_for, request, flash, send_file
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from ...models import db, expr_fecha, Producto, Compra, DetalleCompra, ajustar_stock, registrar_auditoria
from ..permisos import requiere_permiso, requiere_ver_pagina
from ...utils import ahora_bogota, hoy_bogota
from . import admin_bp


@admin_bp.route('/compras')
@login_required
@requiere_ver_pagina('compras')
def compras():
    page    = request.args.get('page', 1, type=int)
    compras = (Compra.query
               .options(
                   db.joinedload(Compra.admin_rel),
                   db.joinedload(Compra.detalles).joinedload(DetalleCompra.producto),
               )
               .order_by(Compra.fechacompra.desc())
               .paginate(page=page, per_page=20, error_out=False))
    productos = Producto.query.filter(Producto.activo.is_(True)).order_by(Producto.nombre).all()
    return render_template('admin/compras.html', compras=compras, productos=productos)


@admin_bp.route('/compras/nueva', methods=['POST'])
@login_required
@requiere_permiso('escribir_todo')
def compra_nueva():
    vendedor   = request.form.get('nombrevendedor', '').strip()
    items      = request.form.getlist('idproducto[]')
    cantidades = request.form.getlist('cantidad[]')
    precios    = request.form.getlist('preciolinea[]')

    if not vendedor or not items:
        flash('Completa todos los campos de la compra.', 'danger')
        return redirect(url_for('admin_panel.compras'))

    total = 0
    detalles = []
    for pid, cant, precio_linea in zip(items, cantidades, precios):
        try:
            pid_int      = int(pid)
            cant_int     = int(cant)
            precio_int   = int(precio_linea)
        except (ValueError, TypeError):
            continue
        prod = Producto.query.get(pid_int)
        if prod and cant_int > 0 and precio_int >= 0:
            # El total de la compra es la suma de los precios pagados por
            # linea, no costo * cantidad: el precio real de una compra
            # puede variar respecto al catalogo.
            total += precio_int
            detalles.append((prod, cant_int, precio_int))

    compra = Compra(
        nombrevendedor = vendedor,
        precio         = total,
        fechacompra    = ahora_bogota(),
        documentoadmin = current_user.documento
    )
    db.session.add(compra)
    db.session.flush()

    cambios_costo = []
    for prod, cant, precio_linea in detalles:
        dc = DetalleCompra(idcompra=compra.idcompra, idproducto=prod.idproducto,
                           cantidad=cant, subtotal=precio_linea)
        db.session.add(dc)
        ajustar_stock(prod.idproducto, cant)
        # Costo unitario efectivo: el catalogo refleja el costo mas reciente.
        costo_nuevo = round(precio_linea / cant)
        if prod.costo != costo_nuevo:
            cambios_costo.append((prod, prod.costo, costo_nuevo))
            prod.costo = costo_nuevo

    db.session.commit()

    registrar_auditoria(current_user.email, 'crear_compra', f'compra:{compra.idcompra}',
                        f'vendedor={vendedor}, total={total}')
    for prod, costo_anterior, costo_nuevo in cambios_costo:
        registrar_auditoria(current_user.email, 'actualizar_costo_por_compra',
                            f'producto:{prod.idproducto}',
                            f'costo {costo_anterior}->{costo_nuevo} (compra #{compra.idcompra})')

    flash('Compra registrada y stock actualizado.', 'success')
    return redirect(url_for('admin_panel.compras'))


@admin_bp.route('/compras/excel')
@login_required
@requiere_ver_pagina('compras')
def compras_excel():
    """Excel de compras de la ultima semana (o del rango ?desde=&hasta=
    si se pasa). Un rango invalido nunca rompe la respuesta."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    MAX_FILAS_EXCEL = 5000
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')
    hoy = hoy_bogota()
    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str \
            else hoy - timedelta(days=6)
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy
    except ValueError:
        desde, hasta = hoy - timedelta(days=6), hoy

    compras_q = (Compra.query
                 .options(db.joinedload(Compra.admin_rel),
                          db.joinedload(Compra.detalles).joinedload(DetalleCompra.producto))
                 .filter(expr_fecha(Compra.fechacompra) >= desde,
                         expr_fecha(Compra.fechacompra) <= hasta)
                 .order_by(Compra.idcompra.asc())
                 .limit(MAX_FILAS_EXCEL)
                 .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    headers = ['ID Compra', 'Vendedor', 'Fecha', 'Hora', 'Total', 'Registrado por', 'Productos']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, c in enumerate(compras_q, 2):
        productos_str = ', '.join(
            f"{d.producto.nombre} x{d.cantidad}" for d in c.detalles if d.producto
        )
        ws.cell(row=row_idx, column=1, value=c.idcompra).border = thin_border
        ws.cell(row=row_idx, column=2, value=c.nombrevendedor).border = thin_border
        ws.cell(row=row_idx, column=3,
                value=c.fechacompra.strftime('%d/%m/%Y') if c.fechacompra else '').border = thin_border
        ws.cell(row=row_idx, column=4,
                value=c.fechacompra.strftime('%H:%M') if c.fechacompra else '').border = thin_border
        ws.cell(row=row_idx, column=5, value=c.precio).border = thin_border
        ws.cell(row=row_idx, column=6,
                value=c.admin_rel.nombre if c.admin_rel else '').border = thin_border
        ws.cell(row=row_idx, column=7, value=productos_str).border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 45

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = ahora_bogota().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f"compras_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
