from flask import render_template, redirect, url_for, request, send_file, flash
from flask_login import login_required
from datetime import datetime, timedelta
from ...models import db, expr_fecha, Venta, DetalleVenta, ajustar_stock
from ..permisos import requiere_permiso, requiere_ver_pagina, puede, iniciales, doc_enmascarado
from ...utils import ahora_bogota, hoy_bogota
from . import admin_bp


@admin_bp.route('/ventas')
@login_required
@requiere_ver_pagina('ventas')
def ventas():
    page = request.args.get('page', 1, type=int)
    periodo = request.args.get('periodo', 'todos')
    hoy = hoy_bogota()

    query = Venta.query.options(
        db.joinedload(Venta.cliente_rel),
        db.joinedload(Venta.detalles).joinedload(DetalleVenta.producto),
    )

    if periodo == 'dia':
        query = query.filter(expr_fecha(Venta.fechaventa) == hoy)
    elif periodo == 'semana':
        desde = hoy - timedelta(days=6)
        query = query.filter(expr_fecha(Venta.fechaventa) >= desde)
    elif periodo == 'mes':
        desde = hoy.replace(day=1)
        query = query.filter(expr_fecha(Venta.fechaventa) >= desde)
    elif periodo == 'anio':
        desde = hoy.replace(month=1, day=1)
        query = query.filter(expr_fecha(Venta.fechaventa) >= desde)

    ventas = query.order_by(Venta.idventa.asc()).paginate(page=page, per_page=15, error_out=False)

    semana_lunes = hoy - timedelta(days=hoy.weekday())
    semana_viernes = semana_lunes + timedelta(days=4)

    return render_template(
        'admin/ventas.html', ventas=ventas, periodo=periodo,
        semana_lunes=semana_lunes.strftime('%Y-%m-%d'),
        semana_viernes=semana_viernes.strftime('%Y-%m-%d'),
    )


def _transicion_venta(idventa, estados_validos, estado_nuevo):
    """UPDATE condicionado al estado actual en BD, para que un doble
    clic no aplique el efecto dos veces."""
    if isinstance(estados_validos, str):
        estados_validos = (estados_validos,)
    afectados = Venta.query.filter(
        Venta.idventa == idventa, Venta.estado.in_(estados_validos)
    ).update({'estado': estado_nuevo}, synchronize_session=False)
    db.session.commit()
    return afectados > 0


@admin_bp.route('/ventas/aceptar/<int:idventa>', methods=['POST'])
@login_required
@requiere_permiso('aceptar_rechazar_venta')
def venta_aceptar(idventa):
    Venta.query.get_or_404(idventa)
    if _transicion_venta(idventa, 'Pendiente de Pago', 'Pagado/Preparando'):
        flash(f'Pedido #{idventa} aceptado y en preparación.', 'success')
    else:
        flash('Solo se pueden aceptar pedidos pendientes.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/rechazar/<int:idventa>', methods=['POST'])
@login_required
@requiere_permiso('aceptar_rechazar_venta')
def venta_rechazar(idventa):
    venta = Venta.query.get_or_404(idventa)
    detalles = list(venta.detalles)  # leer antes de la transicion
    if _transicion_venta(idventa, 'Pendiente de Pago', 'Cancelado'):
        for det in detalles:
            ajustar_stock(det.idproducto, det.cantidad)
        db.session.commit()
        flash(f'Pedido #{idventa} rechazado. Stock devuelto.', 'danger')
    else:
        flash('Solo se pueden rechazar pedidos pendientes.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/preparado/<int:idventa>', methods=['POST'])
@login_required
@requiere_permiso('cambiar_estado_entrega')
def venta_preparado(idventa):
    Venta.query.get_or_404(idventa)
    if _transicion_venta(idventa, 'Pagado/Preparando', 'Preparado'):
        flash(f'Pedido #{idventa} marcado como preparado.', 'success')
    else:
        flash('Solo se pueden preparar pedidos en estado Pagado/Preparando.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/entregado/<int:idventa>', methods=['POST'])
@login_required
@requiere_permiso('cambiar_estado_entrega')
def venta_entregado(idventa):
    Venta.query.get_or_404(idventa)
    if _transicion_venta(idventa, ('Pagado/Preparando', 'Preparado'), 'Entregado'):
        flash(f'Pedido #{idventa} marcado como entregado.', 'success')
    else:
        flash('No se puede entregar este pedido.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/excel')
@login_required
@requiere_ver_pagina('ventas')
def ventas_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    # Rango de fechas opcional (?desde=YYYY-MM-DD&hasta=YYYY-MM-DD); sin
    # parametros, exporta solo el dia de hoy.
    MAX_FILAS_EXCEL = 5000
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')
    hoy = hoy_bogota()
    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else hoy
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy
    except ValueError:
        desde, hasta = hoy, hoy

    ventas = (Venta.query
              .options(db.joinedload(Venta.cliente_rel),
                       db.joinedload(Venta.detalles).joinedload(DetalleVenta.producto))
              .filter(expr_fecha(Venta.fechaventa) >= desde, expr_fecha(Venta.fechaventa) <= hasta)
              .order_by(Venta.idventa.asc())
              .limit(MAX_FILAS_EXCEL)
              .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas del Dia"

    headers = ['#', 'Cliente', 'Documento', 'Ficha', 'Productos', 'Valor Total', 'Estado', 'Fecha', 'Hora']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Sin el permiso ver_datos_personales, exporta datos enmascarados.
    ver_datos = puede('ver_datos_personales')

    for row_idx, v in enumerate(ventas, 2):
        productos_str = ', '.join(
            f"{d.producto.nombre} x{d.cantidad}" for d in v.detalles if d.producto
        )
        nombre_cliente = (v.cliente_rel.nombre if v.cliente_rel else '') if ver_datos \
            else iniciales(v.cliente_rel.nombre if v.cliente_rel else '')
        doc_cliente = v.cliente if ver_datos else doc_enmascarado(v.cliente)
        ficha_cliente = (v.cliente_rel.ficha if v.cliente_rel else '') if ver_datos else '*****'
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=2, value=nombre_cliente).border = thin_border
        ws.cell(row=row_idx, column=3, value=doc_cliente).border = thin_border
        ws.cell(row=row_idx, column=4, value=ficha_cliente).border = thin_border
        ws.cell(row=row_idx, column=5, value=productos_str).border = thin_border
        ws.cell(row=row_idx, column=6, value=v.precio).border = thin_border
        ws.cell(row=row_idx, column=7, value=v.estado).border = thin_border
        ws.cell(row=row_idx, column=8, value=v.fechaventa.strftime('%d/%m/%Y') if v.fechaventa else '').border = thin_border
        ws.cell(row=row_idx, column=9, value=v.fechaventa.strftime('%H:%M') if v.fechaventa else '').border = thin_border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 10

    # Fila de total vendido sobre lo REALMENTE exportado en el rango.
    fila_total = len(ventas) + 2
    fill_total = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    font_total = Font(bold=True)
    totales = {
        5: 'TOTAL VENDIDO',
        6: sum(v.precio for v in ventas),
    }
    for col in range(1, len(headers) + 1):
        celda = ws.cell(row=fila_total, column=col, value=totales.get(col))
        celda.font = font_total
        celda.fill = fill_total
        celda.border = thin_border

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = ahora_bogota().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f"ventas_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
