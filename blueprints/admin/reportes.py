from flask import render_template, redirect, url_for, request, flash, send_file
from flask_login import login_required, current_user
from datetime import datetime
from models import db, Producto, Admin, Reporte
from blueprints.permisos import requiere_permiso, requiere_ver_pagina
from . import admin_bp


# ── Reportes ──────────────────────────────────────────────────────────────────
@admin_bp.route('/reportes')
@login_required
@requiere_ver_pagina('reportes')
def reportes():
    page = request.args.get('page', 1, type=int)
    reportes_q = (Reporte.query
                  .options(db.joinedload(Reporte.prod_rel))
                  .order_by(Reporte.idreporte.desc())
                  .paginate(page=page, per_page=15, error_out=False))
    productos = Producto.query.order_by(Producto.nombre).all()
    return render_template('admin/reportes.html', reportes=reportes_q, productos=productos)


@admin_bp.route('/reportes/crear', methods=['POST'])
@login_required
@requiere_permiso('generar_reporte')
def reporte_crear():
    descripcion = request.form.get('descripcion', '').strip()
    idproducto  = request.form.get('idproducto', '', type=int)

    if not idproducto:
        flash('Debe seleccionar un producto.', 'danger')
        return redirect(url_for('admin_panel.reportes'))

    prod = Producto.query.get(idproducto)
    if not prod:
        flash('Producto no encontrado.', 'danger')
        return redirect(url_for('admin_panel.reportes'))

    uid = current_user.get_id()
    if uid.startswith('admin:'):
        id_creador = current_user.documento
    else:
        id_creador = current_user.docpersonal

    reporte = Reporte(
        idadmin     = id_creador,
        descripcion = descripcion or None,
        fecha       = datetime.utcnow(),
        producto    = idproducto
    )
    db.session.add(reporte)
    db.session.commit()
    flash(f'Reporte #{reporte.idreporte} creado correctamente.', 'success')
    return redirect(url_for('admin_panel.reportes'))


@admin_bp.route('/reportes/excel')
@login_required
@requiere_ver_pagina('reportes')
def reportes_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    # HU-27: rango de fechas opcional (?desde=YYYY-MM-DD&hasta=YYYY-MM-DD)
    # y limite maximo de filas -- antes se exportaba todo el historico sin limite.
    MAX_FILAS_EXCEL = 5000
    query = Reporte.query.options(db.joinedload(Reporte.prod_rel))
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')
    try:
        if desde_str:
            desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
            query = query.filter(Reporte.fecha >= desde)
        if hasta_str:
            hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
            query = query.filter(Reporte.fecha <= hasta)
    except ValueError:
        pass  # rango invalido: se ignora, se exporta sin filtrar por fecha

    reportes = (query
                .order_by(Reporte.idreporte.desc())
                .limit(MAX_FILAS_EXCEL)
                .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = ['ID Reporte', 'ID Admin', 'Nombre Admin', 'Producto', 'Descripción', 'Fecha']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(reportes, 2):
        admin_user = Admin.query.get(r.idadmin)
        ws.cell(row=row_idx, column=1, value=r.idreporte).border = thin_border
        ws.cell(row=row_idx, column=2, value=r.idadmin).border = thin_border
        ws.cell(row=row_idx, column=3, value=admin_user.nombre if admin_user else '').border = thin_border
        ws.cell(row=row_idx, column=4, value=r.prod_rel.nombre if r.prod_rel else '').border = thin_border
        ws.cell(row=row_idx, column=5, value=r.descripcion or '').border = thin_border
        ws.cell(row=row_idx, column=6, value=r.fecha.strftime('%d/%m/%Y') if r.fecha else '').border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f"reportes_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
