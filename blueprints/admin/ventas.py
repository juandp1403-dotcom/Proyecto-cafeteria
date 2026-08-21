from flask import render_template, redirect, url_for, request, send_file, flash
from flask_login import login_required
from datetime import datetime, timedelta
from models import db, Venta, DetalleVenta, ajustar_stock
from blueprints.permisos import requiere_permiso, requiere_ver_pagina
from . import admin_bp


# ── Histórico de ventas ───────────────────────────────────────────────────────
@admin_bp.route('/ventas')
@login_required
@requiere_ver_pagina('ventas')
def ventas():
    from sqlalchemy import cast, Date

    page = request.args.get('page', 1, type=int)
    periodo = request.args.get('periodo', 'todos')
    hoy = datetime.utcnow().date()

    query = Venta.query.options(
        db.joinedload(Venta.cliente_rel),
        db.joinedload(Venta.detalles).joinedload(DetalleVenta.producto),
    )

    if periodo == 'dia':
        query = query.filter(cast(Venta.fechaventa, Date) == hoy)
    elif periodo == 'semana':
        desde = hoy - timedelta(days=6)
        query = query.filter(cast(Venta.fechaventa, Date) >= desde)
    elif periodo == 'mes':
        desde = hoy.replace(day=1)
        query = query.filter(cast(Venta.fechaventa, Date) >= desde)
    elif periodo == 'anio':
        desde = hoy.replace(month=1, day=1)
        query = query.filter(cast(Venta.fechaventa, Date) >= desde)

    ventas = query.order_by(Venta.idventa.asc()).paginate(page=page, per_page=15, error_out=False)

    return render_template('admin/ventas.html', ventas=ventas, periodo=periodo)


def _transicion_venta(idventa, estados_validos, estado_nuevo):
    """HU-48: UPDATE condicionado al estado ACTUAL en la base de datos,
    no al que se leyo minutos antes -- asi un doble clic (dos requests
    casi simultaneos) solo aplica el efecto una vez, la segunda
    peticion ya no encuentra el estado esperado y no hace nada."""
    if isinstance(estados_validos, str):
        estados_validos = (estados_validos,)
    afectados = Venta.query.filter(
        Venta.idventa == idventa, Venta.estado.in_(estados_validos)
    ).update({'estado': estado_nuevo}, synchronize_session=False)
    db.session.commit()
    return afectados > 0


@admin_bp.route('/ventas/aceptar/<int:idventa>', methods=['POST'])
@login_required
@requiere_permiso('aceptar_rechazar_venta')
def venta_aceptar(idventa):
    Venta.query.get_or_404(idventa)
    if _transicion_venta(idventa, 'Pendiente de Pago', 'Pagado/Preparando'):
        flash(f'Pedido #{idventa} aceptado y en preparación.', 'success')
    else:
        flash('Solo se pueden aceptar pedidos pendientes.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/rechazar/<int:idventa>', methods=['POST'])
@login_required
@requiere_permiso('aceptar_rechazar_venta')
def venta_rechazar(idventa):
    venta = Venta.query.get_or_404(idventa)
    detalles = list(venta.detalles)  # leer antes de la transicion
    if _transicion_venta(idventa, 'Pendiente de Pago', 'Cancelado'):
        # HU-47: UPDATE atomico por producto, no lectura+escritura
        for det in detalles:
            ajustar_stock(det.idproducto, det.cantidad)
        db.session.commit()
        flash(f'Pedido #{idventa} rechazado. Stock devuelto.', 'danger')
    else:
        flash('Solo se pueden rechazar pedidos pendientes.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


# ── Cambio de estado en ventas ────────────────────────────────────────────────
@admin_bp.route('/ventas/preparado/<int:idventa>', methods=['POST'])
@login_required
@requiere_permiso('cambiar_estado_entrega')
def venta_preparado(idventa):
    Venta.query.get_or_404(idventa)
    if _transicion_venta(idventa, 'Pagado/Preparando', 'Preparado'):
        flash(f'Pedido #{idventa} marcado como preparado.', 'success')
    else:
        flash('Solo se pueden preparar pedidos en estado Pagado/Preparando.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/entregado/<int:idventa>', methods=['POST'])
@login_required
@requiere_permiso('cambiar_estado_entrega')
def venta_entregado(idventa):
    Venta.query.get_or_404(idventa)
    if _transicion_venta(idventa, ('Pagado/Preparando', 'Preparado'), 'Entregado'):
        flash(f'Pedido #{idventa} marcado como entregado.', 'success')
    else:
        flash('No se puede entregar este pedido.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/excel')
@login_required
@requiere_ver_pagina('ventas')
def ventas_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from sqlalchemy import cast, Date

    # HU-27: rango de fechas opcional (?desde=YYYY-MM-DD&hasta=YYYY-MM-DD).
    # Sin parametros, mantiene el comportamiento original (solo hoy).
    MAX_FILAS_EXCEL = 5000
    desde_str = request.args.get('desde')
    hasta_str = request.args.get('hasta')
    hoy = datetime.utcnow().date()
    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else hoy
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy
    except ValueError:
        desde, hasta = hoy, hoy

    ventas = (Venta.query
              .options(db.joinedload(Venta.cliente_rel),
                       db.joinedload(Venta.detalles).joinedload(DetalleVenta.producto))
              .filter(cast(Venta.fechaventa, Date) >= desde, cast(Venta.fechaventa, Date) <= hasta)
              .order_by(Venta.idventa.asc())
              .limit(MAX_FILAS_EXCEL)
              .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas del Dia"

    headers = ['#', 'Cliente', 'Documento', 'Ficha', 'Productos', 'Valor Total', 'Estado', 'Fecha']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, v in enumerate(ventas, 2):
        productos_str = ', '.join(
            f"{d.producto.nombre} x{d.cantidad}" for d in v.detalles if d.producto
        )
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=2, value=v.cliente_rel.nombre if v.cliente_rel else '').border = thin_border
        ws.cell(row=row_idx, column=3, value=v.cliente).border = thin_border
        ws.cell(row=row_idx, column=4, value=v.cliente_rel.ficha if v.cliente_rel else '').border = thin_border
        ws.cell(row=row_idx, column=5, value=productos_str).border = thin_border
        ws.cell(row=row_idx, column=6, value=v.precio).border = thin_border
        ws.cell(row=row_idx, column=7, value=v.estado).border = thin_border
        ws.cell(row=row_idx, column=8, value=v.fechaventa.strftime('%d/%m/%Y') if v.fechaventa else '').border = thin_border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f"ventas_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
