from flask import render_template, redirect, url_for, request, flash, jsonify, send_file, current_app
from flask_login import login_required, current_user
from datetime import datetime, timedelta
import os
import shutil
import uuid
from models import db, Producto, Admin, Venta, DetalleVenta, Compra, DetalleCompra, BajaInventario
from . import admin_bp


ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def _allowed_image(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def _library_dir():
    return os.path.join(current_app.static_folder, 'imagenes')

def _save_image(file_storage):
    """Guarda imagen en static/productos/ y copia a static/imagenes/ (biblioteca)."""
    if not file_storage or not file_storage.filename:
        return None
    if not _allowed_image(file_storage.filename):
        return None
    ext = file_storage.filename.rsplit('.', 1)[1].lower()
    filename = f"{uuid.uuid4().hex[:12]}.{ext}"
    upload_dir = current_app.config['UPLOAD_FOLDER']
    os.makedirs(upload_dir, exist_ok=True)
    file_storage.save(os.path.join(upload_dir, filename))
    # Copiar a biblioteca para reutilización futura
    lib_dir = _library_dir()
    os.makedirs(lib_dir, exist_ok=True)
    shutil.copy2(os.path.join(upload_dir, filename), os.path.join(lib_dir, filename))
    return filename

def _delete_image(filename):
    """Elimina imagen de static/productos/ (NO de la biblioteca)."""
    if not filename:
        return
    path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(path):
        os.remove(path)

def _get_library_images():
    """Retorna lista de imágenes disponibles en la biblioteca."""
    lib_dir = _library_dir()
    if not os.path.exists(lib_dir):
        return []
    allowed = ALLOWED_EXTENSIONS
    return sorted([
        f for f in os.listdir(lib_dir)
        if '.' in f and f.rsplit('.', 1)[1].lower() in allowed and f != '.gitkeep'
    ])




# ── Dashboard ────────────────────────────────────────────────────────────────
@admin_bp.route('/')
@login_required
def dashboard():
    from sqlalchemy import func, cast, Date, desc, case

    hoy = datetime.utcnow().date()
    estados_pagados = ('Pagado/Preparando', 'Preparado', 'Entregado')

    # KPIs — solo pedidos pagados / entregados
    total_ventas = int(db.session.query(func.coalesce(func.sum(Venta.precio), 0)).filter(
        cast(Venta.fechaventa, Date) == hoy,
        Venta.estado.in_(estados_pagados)
    ).scalar())
    ventas_hoy = int(db.session.query(func.count(Venta.idventa)).filter(
        cast(Venta.fechaventa, Date) == hoy,
        Venta.estado.in_(estados_pagados)
    ).scalar())

    productos_bajo = Producto.query.filter(Producto.stock < 10, Producto.stock > 0).order_by(Producto.stock.asc()).all()
    productos_agotados = Producto.query.filter(Producto.stock == 0).all()

    # ── Ventas diarias (últimos 7 días) — solo pagados/entregados ──
    desde_dias = hoy - timedelta(days=6)
    rows_diarios = db.session.query(
        cast(Venta.fechaventa, Date).label('dia'),
        func.coalesce(func.sum(Venta.precio), 0).label('total')
    ).filter(
        cast(Venta.fechaventa, Date) >= desde_dias,
        Venta.estado.in_(estados_pagados)
    ).group_by('dia').order_by('dia').all()
    mapa_dias = {str(r.dia): int(r.total) for r in rows_diarios}
    ventas_diarias = []
    etiquetas_dias = []
    for i in range(7):
        fecha = desde_dias + timedelta(days=i)
        ventas_diarias.append(mapa_dias.get(str(fecha), 0))
        etiquetas_dias.append(fecha.strftime('%d/%m'))

    # ── Ventas mensuales (últimos 12 meses) — solo pagados/entregados ──
    desde_meses = hoy - timedelta(days=365)
    rows_mensuales = db.session.query(
        func.date_trunc('month', Venta.fechaventa).label('mes'),
        func.coalesce(func.sum(Venta.precio), 0).label('total')
    ).filter(
        cast(Venta.fechaventa, Date) >= desde_meses,
        Venta.estado.in_(estados_pagados)
    ).group_by('mes').order_by('mes').all()
    mapa_meses = {}
    for r in rows_mensuales:
        clave = r.mes.strftime('%Y-%m') if r.mes else ''
        mapa_meses[clave] = int(r.total)
    ventas_mensuales = []
    etiquetas_meses = []
    for i in range(12, 0, -1):
        fecha_mes = hoy.replace(day=1) - timedelta(days=(i - 1) * 30)
        clave = fecha_mes.strftime('%Y-%m')
        ventas_mensuales.append(mapa_meses.get(clave, 0))
        etiquetas_meses.append(fecha_mes.strftime('%b %Y'))

    # ── Top 15 productos más vendidos — solo pagados/entregados ──
    rows_top = (
        db.session.query(
            Producto.nombre,
            func.coalesce(func.sum(DetalleVenta.cantidad), 0).label('total_vendido')
        )
        .join(DetalleVenta, DetalleVenta.idproducto == Producto.idproducto)
        .join(Venta, Venta.idventa == DetalleVenta.idventa)
        .filter(
            cast(Venta.fechaventa, Date) >= hoy - timedelta(days=30),
            Venta.estado.in_(estados_pagados)
        )
        .group_by(Producto.nombre)
        .order_by(desc('total_vendido'))
        .limit(15)
        .all()
    )
    top_15 = [(r.nombre, int(r.total_vendido)) for r in rows_top]

    return render_template('admin/dashboard.html',
                           total_ventas=total_ventas,
                           ventas_hoy=ventas_hoy,
                           productos_bajo=productos_bajo,
                           productos_agotados=productos_agotados,
                           ventas_diarias=ventas_diarias,
                           etiquetas_dias=etiquetas_dias,
                           ventas_mensuales=ventas_mensuales,
                           etiquetas_meses=etiquetas_meses,
                           top_15=top_15)


# ── CRUD Productos ────────────────────────────────────────────────────────────
@admin_bp.route('/productos')
@login_required
def productos():
    prods = Producto.query.order_by(Producto.idproducto).all()
    imagenes = _get_library_images()
    return render_template('admin/productos.html', productos=prods, imagenes_biblioteca=imagenes)


@admin_bp.route('/productos/nuevo', methods=['POST'])
@login_required
def producto_nuevo():
    nombre = request.form.get('nombre', '').strip()
    precio = request.form.get('precio', 0)
    stock  = request.form.get('stock', 0)
    if not nombre:
        flash('El nombre es obligatorio.', 'danger')
        return redirect(url_for('admin_panel.productos'))
    imagen = None
    # Opción 1: subir nueva imagen
    new_image = request.files.get('imagen')
    if new_image and new_image.filename:
        imagen = _save_image(new_image)
    else:
        # Opción 2: seleccionar de biblioteca existente
        lib_image = request.form.get('imagen_biblioteca', '').strip()
        if lib_image:
            # Copiar de biblioteca a static/productos/
            lib_path = os.path.join(_library_dir(), lib_image)
            if os.path.exists(lib_path):
                dest_dir = current_app.config['UPLOAD_FOLDER']
                os.makedirs(dest_dir, exist_ok=True)
                shutil.copy2(lib_path, os.path.join(dest_dir, lib_image))
                imagen = lib_image
    prod = Producto(nombre=nombre, precio=int(precio), stock=int(stock), imagen=imagen)
    db.session.add(prod)
    db.session.commit()
    flash(f'Producto "{nombre}" creado correctamente.', 'success')
    return redirect(url_for('admin_panel.productos'))


@admin_bp.route('/productos/editar/<int:idproducto>', methods=['POST'])
@login_required
def producto_editar(idproducto):
    prod = Producto.query.get_or_404(idproducto)
    prod.nombre = request.form.get('nombre', prod.nombre).strip()
    prod.precio = int(request.form.get('precio', prod.precio))
    prod.stock  = int(request.form.get('stock', prod.stock))
    # Opción 1: subir nueva imagen
    new_image = request.files.get('imagen')
    if new_image and new_image.filename:
        _delete_image(prod.imagen)
        prod.imagen = _save_image(new_image)
    else:
        # Opción 2: seleccionar de biblioteca existente
        lib_image = request.form.get('imagen_biblioteca', '').strip()
        if lib_image:
            lib_path = os.path.join(_library_dir(), lib_image)
            if os.path.exists(lib_path):
                dest_dir = current_app.config['UPLOAD_FOLDER']
                os.makedirs(dest_dir, exist_ok=True)
                shutil.copy2(lib_path, os.path.join(dest_dir, lib_image))
                prod.imagen = lib_image
    db.session.commit()
    flash(f'Producto "{prod.nombre}" actualizado.', 'success')
    return redirect(url_for('admin_panel.productos'))


@admin_bp.route('/productos/eliminar/<int:idproducto>', methods=['POST'])
@login_required
def producto_eliminar(idproducto):
    prod = Producto.query.get_or_404(idproducto)
    _delete_image(prod.imagen)
    db.session.delete(prod)
    db.session.commit()
    flash(f'Producto "{prod.nombre}" eliminado.', 'warning')
    return redirect(url_for('admin_panel.productos'))

@admin_bp.route('/productos/baja/<int:idproducto>', methods=['POST'])
@login_required
def producto_baja(idproducto):
    prod = Producto.query.get_or_404(idproducto)
    cantidad = int(request.form.get('cantidad', 0))
    motivo = request.form.get('motivo', '').strip()
    
    if cantidad <= 0 or not motivo:
        flash('La cantidad y el motivo son obligatorios.', 'danger')
        return redirect(url_for('admin_panel.productos'))
        
    if cantidad > prod.stock:
        flash(f'No puedes dar de baja más unidades de las que hay en stock ({prod.stock}).', 'danger')
        return redirect(url_for('admin_panel.productos'))
        
    prod.stock -= cantidad
    baja = BajaInventario(idproducto=idproducto, cantidad=cantidad, motivo=motivo)
    db.session.add(baja)
    db.session.commit()
    
    flash(f'Se han dado de baja {cantidad} unidades de "{prod.nombre}" por motivo: {motivo}.', 'success')
    return redirect(url_for('admin_panel.productos'))


# ── API: Imágenes de biblioteca ──────────────────────────────────────────────
@admin_bp.route('/productos/imagenes')
@login_required
def productos_imagenes():
    """Retorna JSON con las imágenes disponibles en la biblioteca."""
    imagenes = _get_library_images()
    return jsonify(imagenes)


@admin_bp.route('/productos/excel')
@login_required
def productos_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    productos = Producto.query.order_by(Producto.nombre).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Encabezados
    headers = ['ID', 'Producto', 'Precio ($)', 'Stock', 'Estado']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Datos
    for row_idx, p in enumerate(productos, 2):
        ws.cell(row=row_idx, column=1, value=p.idproducto).border = thin_border
        ws.cell(row=row_idx, column=2, value=p.nombre).border = thin_border
        ws.cell(row=row_idx, column=3, value=p.precio).border = thin_border
        ws.cell(row=row_idx, column=4, value=p.stock).border = thin_border
        ws.cell(row=row_idx, column=5, value=p.estado).border = thin_border

    # Ajustar anchos
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y-%m-%d_%H-%M')
    return send_file(buf, as_attachment=True,
                     download_name=f"inventario_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route('/ventas/excel')
@login_required
def ventas_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from sqlalchemy import cast, Date

    hoy = datetime.utcnow().date()
    ventas = (Venta.query
              .options(db.joinedload(Venta.cliente_rel),
                       db.joinedload(Venta.detalles).joinedload(DetalleVenta.producto))
              .filter(cast(Venta.fechaventa, Date) == hoy)
              .order_by(Venta.idventa.asc())
              .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas del Dia"

    headers = ['#', 'Cliente', 'Documento', 'Ficha', 'Productos', 'Valor Total', 'Estado', 'Fecha']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, v in enumerate(ventas, 2):
        productos_str = ', '.join(
            f"{d.producto.nombre} x{d.cantidad}" for d in v.detalles if d.producto
        )
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=2, value=v.cliente_rel.nombre if v.cliente_rel else '').border = thin_border
        ws.cell(row=row_idx, column=3, value=v.cliente).border = thin_border
        ws.cell(row=row_idx, column=4, value=v.cliente_rel.ficha if v.cliente_rel else '').border = thin_border
        ws.cell(row=row_idx, column=5, value=productos_str).border = thin_border
        ws.cell(row=row_idx, column=6, value=v.precio).border = thin_border
        ws.cell(row=row_idx, column=7, value=v.estado).border = thin_border
        ws.cell(row=row_idx, column=8, value=v.fechaventa.strftime('%d/%m/%Y') if v.fechaventa else '').border = thin_border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f"ventas_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Gestión de usuarios (Admin) ───────────────────────────────────────────────
@admin_bp.route('/usuarios')
@login_required
def usuarios():
    admins = Admin.query.order_by(Admin.documento).all()
    return render_template('admin/usuarios.html', admins=admins)


@admin_bp.route('/usuarios/nuevo', methods=['POST'])
@login_required
def usuario_nuevo():
    doc    = request.form.get('documento', '').strip()
    nombre = request.form.get('nombre', '').strip()
    email  = request.form.get('email', '').strip()
    clave  = request.form.get('clave', '').strip()
    if not all([doc, nombre, email, clave]):
        flash('Todos los campos son obligatorios.', 'danger')
        return redirect(url_for('admin_panel.usuarios'))

    if Admin.query.filter_by(email=email).first():
        flash('Ya existe un usuario con ese correo.', 'danger')
        return redirect(url_for('admin_panel.usuarios'))

    admin = Admin(documento=int(doc), nombre=nombre, email=email, clave='')
    admin.set_password(clave)
    db.session.add(admin)
    db.session.commit()
    flash(f'Usuario "{nombre}" creado correctamente.', 'success')
    return redirect(url_for('admin_panel.usuarios'))


@admin_bp.route('/usuarios/editar/<int:documento>', methods=['POST'])
@login_required
def usuario_editar(documento):
    admin  = Admin.query.get_or_404(documento)
    admin.nombre = request.form.get('nombre', admin.nombre).strip()
    admin.email  = request.form.get('email', admin.email).strip()
    nueva_clave  = request.form.get('clave', '').strip()
    if nueva_clave:
        admin.set_password(nueva_clave)
    db.session.commit()
    flash('Usuario actualizado correctamente.', 'success')
    return redirect(url_for('admin_panel.usuarios'))


@admin_bp.route('/usuarios/eliminar/<int:documento>', methods=['POST'])
@login_required
def usuario_eliminar(documento):
    if documento == current_user.documento:
        flash('No puedes eliminar tu propio usuario.', 'danger')
        return redirect(url_for('admin_panel.usuarios'))
    admin = Admin.query.get_or_404(documento)
    db.session.delete(admin)
    db.session.commit()
    flash('Usuario eliminado.', 'warning')
    return redirect(url_for('admin_panel.usuarios'))


# ── Histórico de ventas ───────────────────────────────────────────────────────
@admin_bp.route('/ventas')
@login_required
def ventas():
    from sqlalchemy import func, cast, Date

    page = request.args.get('page', 1, type=int)
    periodo = request.args.get('periodo', 'todos')
    hoy = datetime.utcnow().date()

    query = Venta.query.options(
        db.joinedload(Venta.cliente_rel),
        db.joinedload(Venta.detalles).joinedload(DetalleVenta.producto),
    )

    if periodo == 'dia':
        query = query.filter(cast(Venta.fechaventa, Date) == hoy)
    elif periodo == 'semana':
        desde = hoy - timedelta(days=6)
        query = query.filter(cast(Venta.fechaventa, Date) >= desde)
    elif periodo == 'mes':
        desde = hoy.replace(day=1)
        query = query.filter(cast(Venta.fechaventa, Date) >= desde)
    elif periodo == 'anio':
        desde = hoy.replace(month=1, day=1)
        query = query.filter(cast(Venta.fechaventa, Date) >= desde)

    ventas = query.order_by(Venta.idventa.asc()).paginate(page=page, per_page=15)

    # Pre-compute order number per day to avoid N+1
    numero_map = {}
    for v in ventas.items:
        key = str(v.fechaventa)
        if key not in numero_map:
            numero_map[key] = db.session.query(func.count(Venta.idventa)).filter(
                Venta.fechaventa == v.fechaventa
            ).scalar()
        v._num_pedido_dia = numero_map[key]

    # Recompute sequential number: within each day, order by id asc
    nums_dia = {}
    for v in ventas.items:
        key = str(v.fechaventa)
        if key not in nums_dia:
            nums_dia[key] = 0
        nums_dia[key] += 1
        v._num_sequential = nums_dia[key]

    return render_template('admin/ventas.html', ventas=ventas, periodo=periodo)


@admin_bp.route('/ventas/aceptar/<int:idventa>', methods=['POST'])
@login_required
def venta_aceptar(idventa):
    venta = Venta.query.get_or_404(idventa)
    if venta.estado == 'Pendiente de Pago':
        venta.estado = 'Pagado/Preparando'
        db.session.commit()
        flash(f'Pedido #{idventa} aceptado y en preparación.', 'success')
    else:
        flash('Solo se pueden aceptar pedidos pendientes.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/rechazar/<int:idventa>', methods=['POST'])
@login_required
def venta_rechazar(idventa):
    venta = Venta.query.get_or_404(idventa)
    if venta.estado == 'Pendiente de Pago':
        # Restore stock
        for det in venta.detalles:
            prod = Producto.query.get(det.idproducto)
            if prod:
                prod.stock += det.cantidad
        venta.estado = 'Cancelado'
        db.session.commit()
        flash(f'Pedido #{idventa} rechazado. Stock devuelto.', 'danger')
    else:
        flash('Solo se pueden rechazar pedidos pendientes.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


# ── Compras / Abastecimiento ──────────────────────────────────────────────────
@admin_bp.route('/compras')
@login_required
def compras():
    page    = request.args.get('page', 1, type=int)
    compras = (Compra.query
               .order_by(Compra.fechacompra.desc())
               .paginate(page=page, per_page=20))
    productos = Producto.query.order_by(Producto.nombre).all()
    return render_template('admin/compras.html', compras=compras, productos=productos)


@admin_bp.route('/compras/nueva', methods=['POST'])
@login_required
def compra_nueva():
    vendedor = request.form.get('nombrevendedor', '').strip()
    items    = request.form.getlist('idproducto[]')
    cantidades = request.form.getlist('cantidad[]')

    if not vendedor or not items:
        flash('Completa todos los campos de la compra.', 'danger')
        return redirect(url_for('admin_panel.compras'))

    total = 0
    detalles = []
    for pid, cant in zip(items, cantidades):
        prod = Producto.query.get(int(pid))
        if prod and int(cant) > 0:
            total += prod.precio * int(cant)
            detalles.append((prod, int(cant)))

    compra = Compra(
        nombrevendedor = vendedor,
        precio         = total,
        fechacompra    = datetime.utcnow(),
        documentoadmin = current_user.documento
    )
    db.session.add(compra)
    db.session.flush()

    for prod, cant in detalles:
        dc = DetalleCompra(idcompra=compra.idcompra, idproducto=prod.idproducto, cantidad=cant)
        db.session.add(dc)
        prod.stock += cant  # suma al inventario

    db.session.commit()
    flash('Compra registrada y stock actualizado.', 'success')
    return redirect(url_for('admin_panel.compras'))


# ── Cambio de estado en ventas ────────────────────────────────────────────────
@admin_bp.route('/ventas/preparado/<int:idventa>', methods=['POST'])
@login_required
def venta_preparado(idventa):
    venta = Venta.query.get_or_404(idventa)
    if venta.estado == 'Pagado/Preparando':
        venta.estado = 'Preparado'
        db.session.commit()
        flash(f'Pedido #{idventa} marcado como preparado.', 'success')
    else:
        flash('Solo se pueden preparar pedidos en estado Pagado/Preparando.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


@admin_bp.route('/ventas/entregado/<int:idventa>', methods=['POST'])
@login_required
def venta_entregado(idventa):
    venta = Venta.query.get_or_404(idventa)
    if venta.estado in ('Pagado/Preparando', 'Preparado'):
        venta.estado = 'Entregado'
        db.session.commit()
        flash(f'Pedido #{idventa} marcado como entregado.', 'success')
    else:
        flash('No se puede entregar este pedido.', 'warning')
    return redirect(url_for('admin_panel.ventas', page=request.args.get('page', 1), periodo=request.args.get('periodo', 'todos')))


# ── Reportes ──────────────────────────────────────────────────────────────────
@admin_bp.route('/reportes')
@login_required
def reportes():
    page = request.args.get('page', 1, type=int)
    reportes_q = (Reporte.query
                  .options(db.joinedload(Reporte.admin_rel),
                           db.joinedload(Reporte.prod_rel))
                  .order_by(Reporte.idreporte.desc())
                  .paginate(page=page, per_page=15))
    productos = Producto.query.order_by(Producto.nombre).all()
    return render_template('admin/reportes.html', reportes=reportes_q, productos=productos)


@admin_bp.route('/reportes/crear', methods=['POST'])
@login_required
def reporte_crear():
    descripcion = request.form.get('descripcion', '').strip()
    idproducto  = request.form.get('idproducto', '', type=int)

    if not idproducto:
        flash('Debe seleccionar un producto.', 'danger')
        return redirect(url_for('admin_panel.reportes'))

    prod = Producto.query.get(idproducto)
    if not prod:
        flash('Producto no encontrado.', 'danger')
        return redirect(url_for('admin_panel.reportes'))

    reporte = Reporte(
        idadmin     = current_user.documento,
        descripcion = descripcion or None,
        fecha       = datetime.utcnow(),
        producto    = idproducto
    )
    db.session.add(reporte)
    db.session.commit()
    flash(f'Reporte #{reporte.idreporte} creado correctamente.', 'success')
    return redirect(url_for('admin_panel.reportes'))


@admin_bp.route('/reportes/excel')
@login_required
def reportes_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    reportes = (Reporte.query
                .options(db.joinedload(Reporte.admin_rel),
                         db.joinedload(Reporte.prod_rel))
                .order_by(Reporte.idreporte.desc())
                .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = ['ID Reporte', 'ID Admin', 'Nombre Admin', 'Producto', 'Descripción', 'Fecha']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(reportes, 2):
        ws.cell(row=row_idx, column=1, value=r.idreporte).border = thin_border
        ws.cell(row=row_idx, column=2, value=r.idadmin).border = thin_border
        ws.cell(row=row_idx, column=3, value=r.admin_rel.nombre if r.admin_rel else '').border = thin_border
        ws.cell(row=row_idx, column=4, value=r.prod_rel.nombre if r.prod_rel else '').border = thin_border
        ws.cell(row=row_idx, column=5, value=r.descripcion or '').border = thin_border
        ws.cell(row=row_idx, column=6, value=r.fecha.strftime('%d/%m/%Y') if r.fecha else '').border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f"reportes_{fecha}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
